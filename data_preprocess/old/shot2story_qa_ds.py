import os
import json
from call_llm import call_llm_completion, call_chatapi_model, call_openai_model, call_silionflow_model, call_vllm_model, call_deepseek_model
from os.path import join, exists
from concurrent.futures import ProcessPoolExecutor
import concurrent
from tqdm import tqdm
import re
import random
import math


prompt_template = """
There is a video, its content is:
{whole_caption}

This video consists of several clips. The content of these clips is: {clip_captions}

Please imagine a scenario: \
Tom asks a question about the video. \
Bob answers the question while watching the video. \
Bob responds immediately after seeing the relevant clips. \
If there are multiple relevant clips, Bob will response for every related clips. \
Bob will supplement his answer with new information when he encounters a new relevant clip.

Now you need to generate the questions and answers in this process.

Please propose a single question that ordinary people would ask. \
The question must can be definitely answered from the clip content. \
Do not generate question that can not be solved with the clip content. \
This question must be related to multiple clips but not all clips. 

Please also give the related clips and answers for the question. \
All your answers constitute a complete answer. \
So do not repeat the contents appeared in previous answers. \
Only include new information in the related clip.

Note that: 
1. The question and answers should not be related to the voice.
2. In the question and answers, treat the video as a single coherent and complete video.\
Do not regard the clips as separate parts.


Please reply in the following json format:
{{
  "question": "the question you ask",
  "related_clips": list of related clip index
  "answers": list of the answers for related clips. Do not mention clip in the answer.
}}
"""

def generate_qa(item, save_dir):
    video = item["video"]
    output_file = join(save_dir, f'{video}.json')
    if exists(output_file):
        print(f'Exists: {output_file}')
        return

    clip_captions = ""
    for i, caption in enumerate(item["captions"]):
        clip_captions = f"{clip_captions}\nClip {i + 1}: {caption}"
    prompt = prompt_template.format(
        whole_caption=item['whole_caption'],
        clip_captions=clip_captions
    )
    messages = [{"role": "user", "content": prompt}]
    print(prompt)

    '''gpt-4o'''
    if MODEL == "gpt-4o":
        result = call_openai_model(messages=messages)
        response, cost = result["response"], result["cost"]

    elif MODEL == "ds_70b":
        response = call_vllm_model(
            messages=messages,
            model='/afs/zengwang/ckpt/DeepSeek-R1-Distill-Llama-70B',
            base_url='http://10.210.0.43:2204/v1'

        )
        result = {"response": response, "cost": 0}

    elif MODEL == "ds_600b":
        # result = call_chatapi_model(messages=messages, model="DeepSeek-R1")
        # response, cost = result["response"], result["cost"]

        # response = call_silionflow_model(messages=messages, model="deepseek-ai/DeepSeek-R1")

        result = call_deepseek_model(messages, model="deepseek-reasoner")
        response, cost = result["response"], result["cost"]

    print(response)
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    return result


def generate_qa_tmp(item, save_dir):
    try:
        generate_qa(item, save_dir)
    except Exception as e:
        print(f"Error processing item {item}: {e}")
    return


def get_clip_frame(clip_name):
    matches = re.findall(r'_(\d+)', clip_name)
    if matches:
        numbers = list(map(int, matches))
    return numbers[-2:]


if __name__ == '__main__':
    # src_file = '/afs/zengwang/projects/task_define_service/data/shot2story/134k_full_train.json'
    # save_dir = '/afs/zengwang/projects/task_define_service/data/shot2story/qas_ds_70b'

    # TEST = True
    TEST = False

    # MODEL = 'ds_600b'
    # MODEL = 'gpt-4o'
    MODEL = 'ds_70b'


    if TEST:
        src_file = '/home/SENSETIME/zengwang/myprojects/task_define_service/data/shot2story/134k_full_train.json'
        save_dir = f'/home/SENSETIME/zengwang/myprojects/task_define_service/data/shot2story/first10/qas_{MODEL}'
        with open(src_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        data = data[:10]
    else:
        src_file = '/home/SENSETIME/zengwang/myprojects/task_define_service/data/shot2story/134k_full_train.json'
        save_dir = f'/home/SENSETIME/zengwang/myprojects/task_define_service/data/shot2story/qas_{MODEL}'
        with open(src_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        if MODEL != 'gpt-4o':
            data = data[10000:]     # 跳过gpt4o的部分

        target_nums = 20 * 1000
        data = data[:target_nums]

    os.makedirs(save_dir, exist_ok=True)

    # max_workers = 30
    # with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
    #     futures = [executor.submit(generate_qa_tmp, item, save_dir) for item in data]
    #     concurrent.futures.wait(futures)

    # for item in data:
    #     generate_qa(item, save_dir)
    #     t = 0

    if TEST:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value='video')
        ws.cell(row=1, column=2, value='whole caption')
        ws.cell(row=1, column=3, value='clip caption')
        ws.cell(row=1, column=4, value='qa')
        ws.cell(row=1, column=5, value='cost')

        for idx, item in enumerate(data, start=2):
            video = item['video']
            ws.cell(row=idx, column=1, value=video)

            whole_caption=item['whole_caption']
            clip_captions = ""
            for i, caption in enumerate(item["captions"]):
                clip_captions = f"{clip_captions}\nClip {i + 1}: {caption}"

            ws.cell(row=idx, column=2, value=whole_caption)
            ws.cell(row=idx, column=3, value=clip_captions)

            result_file = join(save_dir, f"{video}.json")
            if not exists(result_file):
                continue

            with open(result_file, 'r', encoding='utf-8') as f:
                result = json.load(f)
            response = result['response']
            if response is None:
                continue
            if "</think>" in response:
                response = response.split("</think>")[-1]
            response = response.strip()
            if response.startswith('```json'):
                response = response[len("```json"):-len("```")]

            ws.cell(row=idx, column=4, value=response)
            ws.cell(row=idx, column=5, value=result['cost'])

        wb.save(f"/home/SENSETIME/zengwang/myprojects/task_define_service/data/shot2story/first10/qas_{MODEL}.xlsx")

        import sys
        sys.exit(0)


    # # 计算消耗
    # num_all, cost_all = 0, 0
    # for result_file in tqdm(os.listdir(save_dir)):
    #     if not result_file.endswith('.json'):
    #         continue
    #     with open(join(save_dir, result_file), 'r', encoding='utf-8') as f:
    #         result = json.load(f)
    #     cost_all += result['cost']
    #     num_all += 1
    #     # print(result['response'].split('</think>')[-1])
    # print(f'Cost: {cost_all} $ / {num_all} data = {cost_all/num_all} $/data')



    '''汇总QA数据'''
    video_info_file = '/home/SENSETIME/zengwang/myprojects/task_define_service/data/shot2story_video_info.json'
    with open(video_info_file, 'r') as f:
        video_infos = json.load(f)

    # 处理数据
    tar_data = []
    tar_file = f'/home/SENSETIME/zengwang/myprojects/task_define_service/data/shot2story/processed/qa_{MODEL}_v1.json'
    for item in tqdm(data):
        video = item['video']
        result_file = join(save_dir, f"{video}.json")
        if not exists(result_file):
            continue

        with open(result_file, 'r', encoding='utf-8') as f:
            result = json.load(f)
        response = result['response']
        if response is None:
            continue

        valid = True

        response = result['response']
        if response is None:
            continue
        if "</think>" in response:
            response = response.split("</think>")[-1]
        response = response.strip()
        if response.startswith('```json'):
            response = response[len("```json"):-len("```")]

        try:
            video_info = video_infos[join('data', 'shot2story-videos', video)]
            clip_names = item["video_names"]
            clip_frames = [get_clip_frame(clip) for clip in clip_names]
            fps = video_info['frame_rate']
            clip_times = [[st/fps, ed/fps] for st, ed in clip_frames]

            result = json.loads(response)
            question = result["question"]
            related_clips = result["related_clips"]
            answers = result["answers"]

            related_clips, answers = zip(*sorted(zip(related_clips, answers)))
            related_clips = list(related_clips)
            answers = list(answers)
            related_clips = [tmp-1 for tmp in related_clips]

            # 在第一个相关片段前提问
            video_path = join('shot2story-videos', video)
            first_response_time = clip_times[related_clips[0]][1]
            query_time = float(random.randint(0, int(math.floor(first_response_time- 1))))
            messages, videos = [], []
            if query_time > 0:
                messages.append({"role": "user", "content": "<video>", "time": [0.0, query_time]})
                videos.append(video_path)
            messages.append({"role": "user", "content": question, "time": [query_time, query_time]})

            last_time = query_time
            for related_clip, answer in zip(related_clips, answers):
                if not isinstance(answer, str):
                    print(answer)
                    valid = False
                    continue
                response_time = float(clip_times[related_clip][1])
                messages.append({"role": "user", "content": "<video>", "time": [last_time, response_time]})
                videos.append(video_path)
                messages.append({"role": "assistant", "content": answer, "time": [response_time, response_time]})
                last_time = response_time

            # # 剩余的视频如果大于1s，就加入，对于训练stream_head有帮助
            # # 有bug
            # if last_time < video_info['duration'] - 1:
            #     messages.append({"role": "user", "content": "<video>", "time": [last_time,  float(video_info['duration'])]})
            #     videos.append(video_path)

            if valid:
                tar_item = {"messages": messages, "videos": videos}
                tar_data.append(tar_item)

        except:
            print(response)
            continue


    os.makedirs(os.path.dirname(tar_file), exist_ok=True)
    with open(tar_file, 'w', encoding='utf-8') as f:
        json.dump(tar_data, f, ensure_ascii=False)
    print(len(tar_data))
