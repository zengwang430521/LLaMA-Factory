import os
import json
from call_llm import call_llm_completion, call_chatapi_model, call_openai_model, call_silionflow_model, call_vllm_model, call_deepseek_model
from os.path import join, exists
from concurrent.futures import ProcessPoolExecutor
import concurrent
from tqdm import tqdm
import re
import random
import math
from copy import deepcopy


prompt_template = """\
A video consists of multiple clips. Tts content is:
{whole_caption}

The content of these clips is: {clip_captions}.

Your task is to identify individuals who appear in multiple clips and generate a question about them. \
The question should be asked in a clip where the people appear, \
but it must not be answerable based on the information available up to that clip. \
Instead, the answer should only become clear in a later clip where the people appear again.\

Please ensure the individuals envolved in your question and answer is the same.
If there is not individual who appear in multiple clips, \
you can give up this task an response with an empty dict. 


Requirements:
The question and answer must be based on the clip captions.
Do not mention clip in your questions or answers.
The question and answer must not be related to voice or audio content.

Please provide your response in the following JSON format:
{{
  "question_clip": <the clip where the question is asked>,
  "question": "<the generated question>",
  "answer_clip": <the clip where the answer is found>,
  "answer": "<the answer to the question>"
}}

The following is some examples:
{{
  "question_clip": 1,
  "question": "The man whose suit is has deeper color turn his head and look at some, what is he looking at?",
  "answer_clip": 2,
  "answer": "A painting on the wall.",
}}

{{
  "question_clip": 2,
  "question": "The astronauts wake up after a long space journey, what abnormal thing do they find?",
  "answer_clip": 4,
  "answer": "One of the astronauts turns into a strange creature.",
}}

{{
  "question_clip": 4,
  "question": "The man throw the yellow life float into the sea, what action does he take next with it?",
  "answer_clip": 7,
  "answer": "He jumps to the sea and climb to the yellow life float.",
}}
"""

def generate_qa(item, save_dir):
    video = item["video"]
    output_file = join(save_dir, f'{video}.json')
    if exists(output_file):
        print(f'Exists: {output_file}')
        return

    clip_captions = ""
    for i, caption in enumerate(item["captions"]):
        clip_captions = f"{clip_captions}\nClip {i + 1}: {caption}"
    prompt = prompt_template.format(
        whole_caption=item['whole_caption'],
        clip_captions=clip_captions
    )
    messages = [{"role": "user", "content": prompt}]
    print(prompt)

    '''gpt-4o'''
    if MODEL == "gpt-4o":
        result = call_openai_model(messages=messages)
        response, cost = result["response"], result["cost"]

    elif MODEL == "ds_70b":
        response = call_vllm_model(
            messages=messages,
            model='/afs/zengwang/ckpt/DeepSeek-R1-Distill-Llama-70B',
            base_url='http://10.210.0.43:2204/v1'

        )
        result = {"response": response, "cost": 0}

    elif MODEL == "ds_600b":
        # result = call_chatapi_model(messages=messages, model="DeepSeek-R1")
        # response, cost = result["response"], result["cost"]

        # response = call_silionflow_model(messages=messages, model="deepseek-ai/DeepSeek-R1")

        result = call_deepseek_model(messages, model="deepseek-reasoner")
        response, cost = result["response"], result["cost"]

    print(response)
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    return result


def generate_qa_tmp(item, save_dir):
    try:
        generate_qa(item, save_dir)
    except Exception as e:
        print(f"Error processing item {item}: {e}")
    return


def get_clip_frame(clip_name):
    matches = re.findall(r'_(\d+)', clip_name)
    if matches:
        numbers = list(map(int, matches))
    return numbers[-2:]


if __name__ == '__main__':
    # src_file = '/afs/zengwang/projects/task_define_service/data/shot2story/134k_full_train.json'
    # save_dir = '/afs/zengwang/projects/task_define_service/data/shot2story/qas_ds_70b'

    # TEST = True
    TEST = False

    # MODEL = 'ds_600b'
    # MODEL = 'gpt-4o'
    MODEL = 'ds_70b'


    if TEST:
        src_file = '/home/SENSETIME/zengwang/myprojects/task_define_service/data/shot2story/134k_full_train.json'
        save_dir = f'/home/SENSETIME/zengwang/myprojects/task_define_service/data/shot2story/first10/qas_{MODEL}'
        with open(src_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        data = data[:10]
    else:
        src_file = '/home/SENSETIME/zengwang/myprojects/task_define_service/data/shot2story/134k_full_train.json'
        save_dir = f'/home/SENSETIME/zengwang/myprojects/task_define_service/data/shot2story/qas_{MODEL}'
        with open(src_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        # if MODEL != 'gpt-4o':
        #     data = data[10000:]     # 跳过gpt4o的部分
        target_nums = 20 * 1000
        data = data[:target_nums]

    os.makedirs(save_dir, exist_ok=True)

    # max_workers = 100
    # with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
    #     futures = [executor.submit(generate_qa_tmp, item, save_dir) for item in data]
    #     concurrent.futures.wait(futures)

    # for item in data:
    #     generate_qa(item, save_dir)
    #     t = 0

    if TEST:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value='video')
        ws.cell(row=1, column=2, value='whole caption')
        ws.cell(row=1, column=3, value='clip caption')
        ws.cell(row=1, column=4, value='qa')
        ws.cell(row=1, column=5, value='cost')

        for idx, item in enumerate(data, start=2):
            video = item['video']
            ws.cell(row=idx, column=1, value=video)

            whole_caption=item['whole_caption']
            clip_captions = ""
            for i, caption in enumerate(item["captions"]):
                clip_captions = f"{clip_captions}\nClip {i + 1}: {caption}"

            ws.cell(row=idx, column=2, value=whole_caption)
            ws.cell(row=idx, column=3, value=clip_captions)

            result_file = join(save_dir, f"{video}.json")
            if not exists(result_file):
                continue

            with open(result_file, 'r', encoding='utf-8') as f:
                result = json.load(f)
            response = result['response']
            if response is None:
                continue
            if "</think>" in response:
                response = response.split("</think>")[-1]
            response = response.strip()
            if response.startswith('```json'):
                response = response[len("```json"):-len("```")]

            ws.cell(row=idx, column=4, value=response)
            ws.cell(row=idx, column=5, value=result['cost'])

        wb.save(f"/home/SENSETIME/zengwang/myprojects/task_define_service/data/shot2story/first10/qas_{MODEL}.xlsx")

        import sys
        sys.exit(0)


    # # 计算消耗
    # num_all, cost_all = 0, 0
    # for result_file in tqdm(os.listdir(save_dir)):
    #     if not result_file.endswith('.json'):
    #         continue
    #     with open(join(save_dir, result_file), 'r', encoding='utf-8') as f:
    #         result = json.load(f)
    #     cost_all += result['cost']
    #     num_all += 1
    #     # print(result['response'].split('</think>')[-1])
    # print(f'Cost: {cost_all} $ / {num_all} data = {cost_all/num_all} $/data')



    '''汇总QA数据'''
    video_info_file = '/home/SENSETIME/zengwang/myprojects/task_define_service/data/shot2story_video_info.json'
    with open(video_info_file, 'r') as f:
        video_infos = json.load(f)

    # 处理数据
    tar_data = []
    tar_file = f'/home/SENSETIME/zengwang/myprojects/task_define_service/data/shot2story/processed/qa_{MODEL}_v2.json'
    for item in tqdm(data):
        video = item['video']
        result_file = join(save_dir, f"{video}.json")
        if not exists(result_file):
            continue

        with open(result_file, 'r', encoding='utf-8') as f:
            result = json.load(f)
        response = result['response']
        if response is None:
            continue

        valid = True

        response = result['response']
        if response is None:
            continue
        if "</think>" in response:
            response = response.split("</think>")[-1]
        response = response.strip()
        # if response.startswith('```json'):
        #     response = response[len("```json"):-len("```")]
        if '```json' in response:
            response = response.split('```json')[1].split('```')[0].strip()

        try:
            video_info = video_infos[join('data', 'shot2story-videos', video)]
            clip_names = item["video_names"]
            clip_frames = [get_clip_frame(clip) for clip in clip_names]
            fps = video_info['frame_rate']
            clip_times = [[st/fps, ed/fps] for st, ed in clip_frames]

            result = json.loads(response)
            question = result["question"]
            question_clip = result['question_clip']
            answer = result['answer']
            answer_clip = result['answer_clip']
            if question is None:
                valid = False
                continue


            # 在question clip后半段提问
            video_path = join('shot2story-videos', video)
            query_time_range = deepcopy(clip_times[question_clip-1])
            query_time_range[0] = query_time_range[0] * 0.5 + query_time_range[1] * (1-0.5)    # 只在clip的后50%提问
            query_time = random.uniform(query_time_range[0], query_time_range[1])
            query_time = round(query_time)

            '''
            response_period: 表示可以进行回复的区间 [t1, t2, t3, t4]
            0:  不回复
            1:  回复
            -:  不监督
            ......t1 ...... t2 ...... t3 ......t4.......
            000000----------111111111111---------0000000
            '''

            t_start, t_end = clip_times[answer_clip-1]
            delta = t_end - t_start
            response_period = [t_start + 0.4 * delta, t_start + 0.6 * delta, t_end, t_end + 1]
            answer_time = t_end     # 为了充分训练，插入response的位置要尽量靠后一些

            messages, videos = [], []
            if query_time > 0:
                messages.append({"role": "user", "content": "<video>", "time": [0.0, query_time]})
                videos.append(video_path)
            messages.append({"role": "user", "content": question, "time": [query_time, query_time]})


            messages.append({"role": "user", "content": "<video>", "time": [query_time, answer_time]})
            videos.append(video_path)
            messages.append({"role": "assistant", "content": answer, "time": response_period})


            # # 剩余的视频如果大于1s，就加入，对于训练stream_head有帮助
            # # 训练时最后一个message不是assistant时有bug
            # if last_time < video_info['duration'] - 1:
            #     messages.append({"role": "user", "content": "<video>", "time": [last_time,  float(video_info['duration'])]})
            #     videos.append(video_path)

            if valid:
                tar_item = {"messages": messages, "videos": videos}
                tar_data.append(tar_item)

        except:
            print('-'*100)
            print(response)
            continue


    os.makedirs(os.path.dirname(tar_file), exist_ok=True)
    with open(tar_file, 'w', encoding='utf-8') as f:
        json.dump(tar_data, f, ensure_ascii=False)
    print(len(tar_data))
